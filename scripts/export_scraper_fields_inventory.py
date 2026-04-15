from __future__ import annotations

import json
import sqlite3
from pathlib import Path
from typing import Any

import orjson
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "doxx_scraper_fields_for_crm.xlsx"
HEADERS = [
    "field_name",
    "field_path",
    "level",
    "group",
    "data_type",
    "nullable",
    "required_for_publish",
    "required_for_persistence",
    "source_stage",
    "source_location",
    "store_specific_or_common",
    "crm_priority",
    "observed_in_sample",
    "description",
    "example_value",
    "notes",
]


def y(value: bool) -> str:
    return "yes" if value else "no"


def short_json(value: Any, limit: int = 140) -> str:
    text = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return text if len(text) <= limit else text[: limit - 3] + "..."


def make_row(**kwargs: Any) -> dict[str, Any]:
    row = {key: "" for key in HEADERS}
    row.update(kwargs)
    return row


def dict_row(conn: sqlite3.Connection, sql: str, params: tuple[Any, ...] = ()) -> dict[str, Any]:
    raw = conn.execute(sql, params).fetchone()
    if raw is None:
        raise RuntimeError(f"Missing sample row for query: {sql}")
    return dict(raw)


def sqlite_type(decl: str) -> str:
    lowered = (decl or "").lower()
    if "int" in lowered:
        return "integer"
    if "text" in lowered:
        return "string"
    return lowered or "string"


def load_samples() -> dict[str, Any]:
    mediapark_db = ROOT / "data" / "scraper" / "qa" / "mediapark_expanded_v1.db"
    alifshop_db = ROOT / "data" / "scraper" / "qa" / "alifshop.db"
    with sqlite3.connect(mediapark_db) as conn:
        conn.row_factory = sqlite3.Row
        scrape_run = dict_row(conn, "select * from scrape_runs order by id asc limit 1")
        raw_product = dict_row(conn, "select * from raw_products order by id asc limit 1")
        outbox = dict_row(conn, "select * from publication_outbox order by id asc limit 1")
        image_row = dict_row(
            conn,
            "select * from raw_product_images where raw_product_id = ? order by position asc limit 1",
            (raw_product["id"],),
        )
        spec_row = dict_row(
            conn,
            "select * from raw_product_specs where raw_product_id = ? order by position asc limit 1",
            (raw_product["id"],),
        )
    with sqlite3.connect(alifshop_db) as conn:
        conn.row_factory = sqlite3.Row
        attempt_row = dict_row(conn, "select * from publication_attempts order by id asc limit 1")
        published_outbox = dict_row(
            conn,
            "select * from publication_outbox where status = 'published' order by id asc limit 1",
        )

    structured = orjson.loads(raw_product["structured_payload_json"])
    raw_payload = orjson.loads(raw_product["raw_payload_json"])
    outbox_payload = orjson.loads(outbox["payload_json"])
    publication = outbox_payload["publication"]
    raw_snapshot = structured["raw_payload_snapshot"]
    run_stats = orjson.loads(scrape_run["stats_json"])
    run_categories = orjson.loads(scrape_run["category_urls_json"])

    raw_key_union: set[str] = set()
    for db_name in (
        "mediapark_expanded_v1.db",
        "uzum_expanded_v1.db",
        "texnomart_expanded_v1.db",
        "alifshop_expanded_v1.db",
    ):
        path = ROOT / "data" / "scraper" / "qa" / db_name
        with sqlite3.connect(path) as conn:
            conn.row_factory = sqlite3.Row
            for raw in conn.execute("select raw_payload_json from raw_products limit 10"):
                raw_key_union.update(orjson.loads(raw["raw_payload_json"]).keys())

    return {
        "scrape_run": scrape_run,
        "raw_product": raw_product,
        "outbox": outbox,
        "image_row": image_row,
        "spec_row": spec_row,
        "attempt_row": attempt_row,
        "published_outbox": published_outbox,
        "structured": structured,
        "raw_payload": raw_payload,
        "outbox_payload": outbox_payload,
        "publication": publication,
        "raw_snapshot": raw_snapshot,
        "run_stats": run_stats,
        "run_categories": run_categories,
        "raw_key_union": sorted(raw_key_union),
    }


def write_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for item in rows:
        ws.append([item.get(header, "") for header in headers])
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wrap_cols = {"field_path", "source_location", "description", "example_value", "notes"}
    for index, header in enumerate(headers, start=1):
        values = [str(header)] + [str(item.get(header, "")) for item in rows]
        ws.column_dimensions[get_column_letter(index)].width = min(max(len(max(values, key=len)) + 2, 12), 60)
        for cell in ws[get_column_letter(index)]:
            cell.alignment = Alignment(vertical="top", wrap_text=header in wrap_cols)


def build_contract_rows(sample: dict[str, Any]) -> list[dict[str, Any]]:
    structured = sample["structured"]
    raw_snapshot = sample["raw_snapshot"]
    payload = sample["outbox_payload"]
    publication = sample["publication"]
    spec_row = sample["spec_row"]
    image_row = sample["image_row"]

    meta = {
        "event_id": ("message_contract", "string", "outbox payload", "domain/publication_event.py::ScraperProductEvent.event_id", "common", "key", "Stable event identifier.", "CRM should dedupe by this first."),
        "event_type": ("message_contract", "string", "outbox payload", "config/settings.py::SCRAPER_OUTBOX_EVENT_TYPE", "common", "key", "Logical event type.", "Current constant: scraper.product.scraped.v1."),
        "schema_version": ("message_contract", "integer", "outbox payload", "config/settings.py::MESSAGE_SCHEMA_VERSION", "common", "key", "Schema version of the event body.", "Current version is 1."),
        "scrape_run_id": ("metadata", "string", "pipeline", "infrastructure/pipelines/scraper_storage_pipeline.py::open_spider", "common", "diagnostic", "Scrape run identifier.", "Useful for replay and QA slicing."),
        "store_name": ("metadata", "string", "shaping", "domain/scraped_product.py::ScrapedProductSnapshot.store_name", "common", "key", "Store name.", "Observed values: mediapark, texnomart, uzum, alifshop."),
        "source_id": ("metadata", "string", "spider", "infrastructure/spiders/*::extract_source_id_from_url", "common", "key", "Store-native product id, if extracted.", "Nullable by contract."),
        "source_url": ("metadata", "string", "spider", "domain/scraped_product.py::from_scrapy_item", "common", "key", "Source PDP URL.", "Used in identity fallback when source_id is missing."),
        "scraped_at": ("metadata", "datetime", "shaping", "domain/scraped_product.py::ScrapedProductSnapshot.scraped_at", "common", "recommended", "UTC timestamp when snapshot was built.", "ISO-8601 / Z format."),
        "payload_hash": ("metadata", "string", "shaping", "domain/scrape_fingerprints.py::build_scraped_payload_hash", "common", "key", "Business fingerprint of minimally structured payload.", "Use with identity_key for business dedupe."),
        "structured_payload": ("message_contract", "object", "outbox payload", "domain/publication_event.py::ScrapedProductPayload", "common", "key", "Main business payload for CRM.", "Contains the minimally structured raw product snapshot."),
        "publication": ("outbox", "object", "outbox payload", "domain/publication_event.py::PublicationMetadata", "common", "diagnostic", "Publication metadata block.", "Diagnostic only; not product business data."),
        "structured_payload.store_name": ("metadata", "string", "shaping", "domain/publication_event.py::ScrapedProductPayload.store_name", "common", "key", "Store name repeated inside payload.", ""),
        "structured_payload.source_url": ("metadata", "string", "spider", "domain/publication_event.py::ScrapedProductPayload.source_url", "common", "key", "Source URL repeated inside payload.", ""),
        "structured_payload.source_id": ("metadata", "string", "spider", "domain/publication_event.py::ScrapedProductPayload.source_id", "common", "key", "Source id inside payload.", "Nullable by contract."),
        "structured_payload.title": ("core_product", "string", "spider", "domain/publication_event.py::ScrapedProductPayload.title", "common", "key", "Product title after minimal cleanup.", "ValidatePipeline drops items without title/name."),
        "structured_payload.brand": ("core_product", "string", "spider", "domain/publication_event.py::ScrapedProductPayload.brand", "common", "recommended", "Brand if available.", ""),
        "structured_payload.price_raw": ("core_product", "string", "spider", "domain/publication_event.py::ScrapedProductPayload.price_raw", "common", "key", "Raw store price string.", "No numeric normalization happens in scraper contour."),
        "structured_payload.in_stock": ("core_product", "boolean", "validation", "domain/publication_event.py::ScrapedProductPayload.in_stock", "common", "recommended", "Availability flag after minimal coercion.", ""),
        "structured_payload.raw_specs": ("specs", "object", "spider", "domain/publication_event.py::ScrapedProductPayload.raw_specs", "common", "key", "Loss-minimizing raw specs map.", "Container required by contract; keys are dynamic."),
        "structured_payload.image_urls": ("images", "array", "spider", "domain/publication_event.py::ScrapedProductPayload.image_urls", "common", "key", "Ordered product image URL list.", "Container required by contract."),
        "structured_payload.description": ("core_product", "string", "spider", "domain/publication_event.py::ScrapedProductPayload.description", "common", "recommended", "Store-native description if available.", ""),
        "structured_payload.category_hint": ("core_product", "string", "shaping", "domain/publication_event.py::ScrapedProductPayload.category_hint", "common", "recommended", "Lightweight category hint.", "Not a final taxonomy mapping."),
        "structured_payload.external_ids": ("metadata", "object", "shaping", "domain/publication_event.py::ScrapedProductPayload.external_ids", "common", "recommended", "External ids map.", "Usually includes {store_name: source_id}."),
        "structured_payload.scraped_at": ("metadata", "datetime", "shaping", "domain/publication_event.py::ScrapedProductPayload.scraped_at", "common", "diagnostic", "Repeat of top-level scraped_at.", ""),
        "structured_payload.payload_hash": ("metadata", "string", "shaping", "domain/publication_event.py::ScrapedProductPayload.payload_hash", "common", "key", "Repeat of top-level payload_hash.", ""),
        "structured_payload.raw_payload_snapshot": ("raw_payload", "object", "spider", "domain/publication_event.py::ScrapedProductPayload.raw_payload_snapshot", "common", "diagnostic", "Validated raw spider item snapshot.", "Keys are not schema-enforced individually."),
        "structured_payload.scrape_run_id": ("metadata", "string", "pipeline", "domain/publication_event.py::ScrapedProductPayload.scrape_run_id", "common", "diagnostic", "Repeat of run id inside payload.", ""),
        "structured_payload.identity_key": ("metadata", "string", "shaping", "domain/scrape_fingerprints.py::build_product_identity_key", "common", "key", "Stable source-product identity key.", "Best business key for downstream mapping."),
        "publication.publication_version": ("outbox", "integer", "outbox payload", "domain/publication_event.py::PublicationMetadata.publication_version", "common", "diagnostic", "Publication metadata version.", ""),
        "publication.exchange_name": ("outbox", "string", "outbox payload", "application/ingestion/event_builder.py::build_scraper_product_event", "common", "diagnostic", "Configured exchange name.", "Config-derived."),
        "publication.queue_name": ("outbox", "string", "outbox payload", "application/ingestion/event_builder.py::build_scraper_product_event", "common", "diagnostic", "Configured queue name embedded in payload metadata.", "Config-derived; not a dedicated DB column."),
        "publication.routing_key": ("outbox", "string", "outbox payload", "application/ingestion/event_builder.py::build_scraper_product_event", "common", "diagnostic", "Configured routing key.", "Config-derived."),
        "publication.outbox_status": ("outbox", "string", "outbox payload", "domain/publication_event.py::PublicationMetadata.outbox_status", "common", "diagnostic", "Outbox status captured in payload metadata.", "Rabbit publish uses a final event body with outbox_status='published'."),
        "publication.attempt_number": ("outbox", "integer", "outbox payload", "domain/publication_event.py::PublicationMetadata.attempt_number", "common", "diagnostic", "Attempt number captured in payload metadata.", "Actual attempts are audited in publication_attempts table."),
        "publication.publisher_service": ("outbox", "string", "outbox payload", "domain/publication_event.py::PublicationMetadata.publisher_service", "common", "diagnostic", "Publisher service name if populated.", "Rabbit publish requires a non-empty publisher_service string."),
        "publication.outbox_created_at": ("outbox", "datetime", "outbox payload", "domain/publication_event.py::PublicationMetadata.outbox_created_at", "common", "diagnostic", "Timestamp when outbox row was created.", ""),
        "publication.published_at": ("outbox", "datetime", "outbox payload", "domain/publication_event.py::PublicationMetadata.published_at", "common", "diagnostic", "Timestamp when publisher marked message as published, if populated.", "Rabbit publish requires published_at to be set on the final event body."),
    }
    required_publish = {
        "event_id", "event_type", "schema_version", "scrape_run_id", "store_name", "source_url", "scraped_at", "payload_hash", "structured_payload", "publication",
        "structured_payload.store_name", "structured_payload.source_url", "structured_payload.title", "structured_payload.raw_specs", "structured_payload.image_urls", "structured_payload.scraped_at", "structured_payload.payload_hash", "structured_payload.raw_payload_snapshot", "structured_payload.scrape_run_id", "structured_payload.identity_key",
        "publication.publication_version", "publication.outbox_status", "publication.outbox_created_at",
    }
    required_persist = {
        "event_id", "event_type", "schema_version", "scrape_run_id", "store_name", "source_url", "scraped_at", "payload_hash", "structured_payload", "publication",
        "structured_payload.store_name", "structured_payload.source_url", "structured_payload.title", "structured_payload.raw_specs", "structured_payload.image_urls", "structured_payload.scraped_at", "structured_payload.payload_hash", "structured_payload.raw_payload_snapshot", "structured_payload.scrape_run_id", "structured_payload.identity_key",
        "publication.publication_version", "publication.exchange_name", "publication.queue_name", "publication.routing_key", "publication.outbox_status", "publication.attempt_number", "publication.outbox_created_at",
    }
    examples = {
        "event_id": payload["event_id"],
        "event_type": payload["event_type"],
        "schema_version": payload["schema_version"],
        "scrape_run_id": payload["scrape_run_id"],
        "store_name": payload["store_name"],
        "source_id": payload["source_id"],
        "source_url": payload["source_url"],
        "scraped_at": payload["scraped_at"],
        "payload_hash": payload["payload_hash"],
        "structured_payload": short_json({"title": structured["title"], "price_raw": structured["price_raw"], "identity_key": structured["identity_key"]}),
        "publication": short_json(publication),
        "structured_payload.store_name": structured["store_name"],
        "structured_payload.source_url": structured["source_url"],
        "structured_payload.source_id": structured["source_id"],
        "structured_payload.title": structured["title"],
        "structured_payload.brand": structured["brand"],
        "structured_payload.price_raw": structured["price_raw"],
        "structured_payload.in_stock": str(structured["in_stock"]).lower(),
        "structured_payload.raw_specs": short_json(structured["raw_specs"]),
        "structured_payload.image_urls": short_json(structured["image_urls"][:2]),
        "structured_payload.description": structured["description"],
        "structured_payload.category_hint": structured["category_hint"],
        "structured_payload.external_ids": short_json(structured["external_ids"]),
        "structured_payload.scraped_at": structured["scraped_at"],
        "structured_payload.payload_hash": structured["payload_hash"],
        "structured_payload.raw_payload_snapshot": short_json(raw_snapshot),
        "structured_payload.scrape_run_id": structured["scrape_run_id"],
        "structured_payload.identity_key": structured["identity_key"],
        "publication.publication_version": publication["publication_version"],
        "publication.exchange_name": publication["exchange_name"],
        "publication.queue_name": publication["queue_name"],
        "publication.routing_key": publication["routing_key"],
        "publication.outbox_status": publication["outbox_status"],
        "publication.attempt_number": publication["attempt_number"],
        "publication.publisher_service": publication["publisher_service"],
        "publication.outbox_created_at": publication["outbox_created_at"],
        "publication.published_at": publication["published_at"],
    }
    nullable = {
        "source_id", "structured_payload.source_id", "structured_payload.brand", "structured_payload.price_raw", "structured_payload.in_stock",
        "structured_payload.description", "structured_payload.category_hint", "publication.exchange_name", "publication.queue_name",
        "publication.routing_key", "publication.publisher_service", "publication.published_at",
    }

    rows: list[dict[str, Any]] = []
    for path, values in meta.items():
        group, dtype, stage, location, scope, priority, desc, notes = values
        rows.append(
            make_row(
                field_name=path.split(".")[-1],
                field_path=path,
                level="top_level" if "." not in path else "nested_object",
                group=group,
                data_type=dtype,
                nullable=y(path in nullable),
                required_for_publish=y(path in required_publish),
                required_for_persistence=y(path in required_persist),
                source_stage=stage,
                source_location=location,
                store_specific_or_common=scope,
                crm_priority=priority,
                observed_in_sample="yes",
                description=desc,
                example_value=str(examples.get(path, "")),
                notes=notes,
            )
        )

    rows.extend(
        [
            make_row(field_name="raw_specs_entry", field_path="structured_payload.raw_specs.<dynamic_key>", level="array_item", group="specs", data_type="dynamic key/value", nullable="yes", required_for_publish="no", required_for_persistence="no", source_stage="spider", source_location="infrastructure/spiders/*::full_parse_item; domain/publication_event.py::ScrapedProductPayload.raw_specs", store_specific_or_common="store-specific", crm_priority="key", observed_in_sample="yes", description="One spec label/value pair inside raw_specs.", example_value=f"{spec_row['spec_name']} => {spec_row['spec_value']}", notes="Dynamic labels vary by store and category; do not assume a fixed key set."),
            make_row(field_name="image_url_item", field_path="structured_payload.image_urls[]", level="array_item", group="images", data_type="string", nullable="no", required_for_publish="no", required_for_persistence="no", source_stage="spider", source_location="infrastructure/spiders/*::full_parse_item; domain/publication_event.py::ScrapedProductPayload.image_urls", store_specific_or_common="common", crm_priority="recommended", observed_in_sample="yes", description="One image URL inside image_urls array.", example_value=str(image_row["image_url"]), notes="Image position is preserved separately in raw_product_images."),
            make_row(field_name="external_id_entry", field_path="structured_payload.external_ids.<store_name>", level="array_item", group="metadata", data_type="dynamic key/value", nullable="yes", required_for_publish="no", required_for_persistence="no", source_stage="shaping", source_location="domain/scraped_product.py::from_scrapy_item", store_specific_or_common="store-specific", crm_priority="recommended", observed_in_sample="yes", description="One external_ids entry keyed by source system or store name.", example_value="mediapark => 27001", notes="Observed live keys: mediapark, texnomart, uzum, alifshop."),
            make_row(field_name="raw_specs_entry", field_path="structured_payload.raw_payload_snapshot.raw_specs.<dynamic_key>", level="array_item", group="raw_payload", data_type="dynamic key/value", nullable="yes", required_for_publish="no", required_for_persistence="no", source_stage="spider", source_location="domain/raw_product.py::RawProduct.raw_specs", store_specific_or_common="store-specific", crm_priority="diagnostic", observed_in_sample="yes", description="One dynamic raw spec inside raw payload snapshot.", example_value=f"{spec_row['spec_name']} => {spec_row['spec_value']}", notes="Raw snapshot is audit/debug data, not final business normalization."),
            make_row(field_name="image_url_item", field_path="structured_payload.raw_payload_snapshot.image_urls[]", level="array_item", group="raw_payload", data_type="string", nullable="no", required_for_publish="no", required_for_persistence="no", source_stage="spider", source_location="domain/raw_product.py::RawProduct.image_urls", store_specific_or_common="common", crm_priority="diagnostic", observed_in_sample="yes", description="One image URL inside raw payload snapshot.", example_value=str(image_row["image_url"]), notes="Final business image array lives in structured_payload.image_urls."),
            make_row(field_name="external_id_entry", field_path="structured_payload.raw_payload_snapshot.external_ids.<store_name>", level="array_item", group="raw_payload", data_type="dynamic key/value", nullable="yes", required_for_publish="no", required_for_persistence="no", source_stage="spider", source_location="domain/raw_product.py::RawProduct.external_ids", store_specific_or_common="store-specific", crm_priority="diagnostic", observed_in_sample="yes", description="One raw external_ids entry inside raw payload snapshot.", example_value="mediapark => 27001", notes="Usually mirrors structured_payload.external_ids."),
        ]
    )
    for key in sorted(raw_snapshot.keys()):
        value = raw_snapshot[key]
        rows.append(
            make_row(
                field_name=key,
                field_path=f"structured_payload.raw_payload_snapshot.{key}",
                level="nested_object" if isinstance(value, (dict, list)) else "top_level",
                group="raw_payload",
                data_type="object" if isinstance(value, dict) else "array" if isinstance(value, list) else "boolean" if isinstance(value, bool) else "string",
                nullable="yes",
                required_for_publish="no",
                required_for_persistence="no",
                source_stage="spider",
                source_location="domain/raw_product.py::RawProduct; domain/scraped_product.py::from_scrapy_item",
                store_specific_or_common="common",
                crm_priority="diagnostic",
                observed_in_sample="yes",
                description=f"Raw payload snapshot field {key}.",
                example_value=short_json(value) if isinstance(value, (dict, list)) else str(value),
                notes="Observed union of raw payload keys in live DBs: " + ", ".join(sample["raw_key_union"]),
            )
        )
    return rows


def build_db_rows(sample: dict[str, Any]) -> list[dict[str, Any]]:
    mediapark_db = ROOT / "data" / "scraper" / "qa" / "mediapark_expanded_v1.db"
    scrape_run = sample["scrape_run"]
    raw_product = sample["raw_product"]
    image_row = sample["image_row"]
    spec_row = sample["spec_row"]
    published_outbox = sample["published_outbox"]
    attempt_row = sample["attempt_row"]
    run_categories = sample["run_categories"]

    sample_rows = {
        "scrape_runs": scrape_run,
        "raw_products": raw_product,
        "raw_product_images": image_row,
        "raw_product_specs": spec_row,
        "publication_outbox": published_outbox,
        "publication_attempts": attempt_row,
    }
    stage_map = {
        "scrape_runs": "pipeline",
        "raw_products": "db persistence",
        "raw_product_images": "db persistence",
        "raw_product_specs": "db persistence",
        "publication_outbox": "db persistence",
        "publication_attempts": "publisher",
    }
    desc_map = {
        "scrape_runs.id": ("Run row surrogate id.", "db_only"),
        "scrape_runs.run_id": ("Stable scraper run identifier.", "diagnostic"),
        "scrape_runs.store_name": ("Store name for the run.", "diagnostic"),
        "scrape_runs.spider_name": ("Scrapy spider name.", "db_only"),
        "scrape_runs.started_at": ("Run start timestamp.", "diagnostic"),
        "scrape_runs.finished_at": ("Run finish timestamp.", "diagnostic"),
        "scrape_runs.status": ("Run status.", "diagnostic"),
        "scrape_runs.items_scraped": ("Items yielded by spider in this run.", "diagnostic"),
        "scrape_runs.items_persisted": ("Items persisted into scraper DB.", "diagnostic"),
        "scrape_runs.items_failed": ("Failed/dropped items in run.", "diagnostic"),
        "scrape_runs.category_urls_json": ("JSON array of seeded category URLs.", "diagnostic"),
        "scrape_runs.stats_json": ("JSON map of run-level stats and coverage counters.", "diagnostic"),
        "scrape_runs.created_at": ("Row creation timestamp.", "db_only"),
        "scrape_runs.updated_at": ("Row update timestamp.", "db_only"),
        "raw_products.id": ("Raw product row surrogate id.", "db_only"),
        "raw_products.scrape_run_id": ("Run id for this product snapshot.", "diagnostic"),
        "raw_products.store_name": ("Store name copied into raw_products.", "key"),
        "raw_products.source_id": ("Store-native product id if available.", "key"),
        "raw_products.source_url": ("Source PDP URL.", "key"),
        "raw_products.identity_key": ("Stable source-product identity key.", "key"),
        "raw_products.title": ("Persisted minimally structured title.", "key"),
        "raw_products.brand": ("Persisted brand.", "recommended"),
        "raw_products.price_raw": ("Persisted raw price string.", "key"),
        "raw_products.in_stock": ("Persisted stock flag stored as nullable integer 0/1.", "recommended"),
        "raw_products.description": ("Persisted description.", "recommended"),
        "raw_products.category_hint": ("Persisted lightweight category hint.", "recommended"),
        "raw_products.external_ids_json": ("JSON map of external ids.", "recommended"),
        "raw_products.payload_hash": ("Persisted business fingerprint.", "key"),
        "raw_products.raw_payload_json": ("JSON snapshot of validated raw spider item.", "diagnostic"),
        "raw_products.structured_payload_json": ("Full shaped payload stored as JSON.", "diagnostic"),
        "raw_products.scraped_at": ("UTC scrape timestamp.", "diagnostic"),
        "raw_products.publication_state": ("Current publish lifecycle state mirrored from outbox.", "diagnostic"),
        "raw_products.created_at": ("Row creation timestamp.", "db_only"),
        "raw_products.updated_at": ("Row update timestamp.", "db_only"),
        "raw_product_images.id": ("Image row surrogate id.", "db_only"),
        "raw_product_images.raw_product_id": ("FK to raw_products.id.", "db_only"),
        "raw_product_images.image_url": ("One image URL stored as a first-class row.", "recommended"),
        "raw_product_images.position": ("Ordinal position of image within product image array.", "recommended"),
        "raw_product_images.created_at": ("Row creation timestamp.", "db_only"),
        "raw_product_images.updated_at": ("Row update timestamp.", "db_only"),
        "raw_product_specs.id": ("Spec row surrogate id.", "db_only"),
        "raw_product_specs.raw_product_id": ("FK to raw_products.id.", "db_only"),
        "raw_product_specs.spec_name": ("Flattened raw spec name.", "recommended"),
        "raw_product_specs.spec_value": ("Flattened raw spec value.", "recommended"),
        "raw_product_specs.source_section": ("Optional parent section when raw_specs had nested objects.", "diagnostic"),
        "raw_product_specs.position": ("Ordinal position of flattened spec row.", "diagnostic"),
        "raw_product_specs.created_at": ("Row creation timestamp.", "db_only"),
        "raw_product_specs.updated_at": ("Row update timestamp.", "db_only"),
        "publication_outbox.id": ("Outbox row surrogate id.", "db_only"),
        "publication_outbox.raw_product_id": ("Unique link back to raw_products.id.", "db_only"),
        "publication_outbox.event_id": ("Stable outbox event id.", "key"),
        "publication_outbox.event_type": ("Event type stored with outbox row.", "diagnostic"),
        "publication_outbox.schema_version": ("Schema version stored with outbox row.", "diagnostic"),
        "publication_outbox.scrape_run_id": ("Run id attached to outbox row.", "diagnostic"),
        "publication_outbox.store_name": ("Store name attached to outbox row.", "diagnostic"),
        "publication_outbox.source_id": ("Optional source id copied into outbox row.", "diagnostic"),
        "publication_outbox.source_url": ("Source PDP URL copied into outbox row.", "diagnostic"),
        "publication_outbox.payload_hash": ("Business fingerprint copied into outbox row.", "key"),
        "publication_outbox.exchange_name": ("Exchange used by publisher topology.", "diagnostic"),
        "publication_outbox.routing_key": ("Routing key used for publication.", "diagnostic"),
        "publication_outbox.payload_json": ("Serialized ScraperProductEvent body.", "diagnostic"),
        "publication_outbox.status": ("Current outbox status.", "diagnostic"),
        "publication_outbox.available_at": ("When the outbox row is eligible for claim/retry.", "diagnostic"),
        "publication_outbox.published_at": ("When publisher marked row as published.", "diagnostic"),
        "publication_outbox.retry_count": ("Attempt count already consumed by this outbox row.", "diagnostic"),
        "publication_outbox.last_error": ("Last publisher error message, if any.", "diagnostic"),
        "publication_outbox.lease_owner": ("Current publisher lease owner.", "diagnostic"),
        "publication_outbox.lease_expires_at": ("Lease expiry timestamp.", "diagnostic"),
        "publication_outbox.created_at": ("Outbox row creation timestamp.", "db_only"),
        "publication_outbox.updated_at": ("Outbox row update timestamp.", "db_only"),
        "publication_attempts.id": ("Publication attempt row surrogate id.", "db_only"),
        "publication_attempts.outbox_id": ("FK to publication_outbox.id.", "db_only"),
        "publication_attempts.attempt_number": ("1-based attempt number per outbox row.", "diagnostic"),
        "publication_attempts.attempted_at": ("Timestamp when publisher attempted delivery.", "diagnostic"),
        "publication_attempts.success": ("Whether the attempt succeeded (1) or failed (0).", "diagnostic"),
        "publication_attempts.error_message": ("Publisher error text for failed attempts.", "diagnostic"),
        "publication_attempts.publisher_name": ("Publisher service identifier.", "diagnostic"),
        "publication_attempts.created_at": ("Attempt row creation timestamp.", "db_only"),
    }
    nullable_override = {
        "scrape_runs.finished_at", "raw_products.source_id", "raw_products.brand", "raw_products.price_raw", "raw_products.in_stock",
        "raw_products.description", "raw_products.category_hint", "raw_product_specs.source_section", "publication_outbox.source_id",
        "publication_outbox.published_at", "publication_outbox.last_error", "publication_outbox.lease_owner",
        "publication_outbox.lease_expires_at", "publication_attempts.error_message", "publication_attempts.publisher_name",
    }
    notes_map = {
        "raw_products.external_ids_json": "Concrete nested shape mirrors structured_payload.external_ids.",
        "raw_products.raw_payload_json": "Nested live keys are documented on publish_contract sheet under structured_payload.raw_payload_snapshot.*.",
        "raw_products.structured_payload_json": "Nested live keys are documented on publish_contract sheet under structured_payload.*.",
        "publication_outbox.payload_json": "Contains the same event contract documented on publish_contract sheet.",
        "raw_product_specs.spec_name": "Labels are intentionally store-native and not normalized here.",
        "raw_product_specs.spec_value": "Flattened from raw_specs for SQL/XLSX-friendly export.",
        "publication_attempts.success": "Stored as integer 0/1 in SQLite.",
    }

    rows: list[dict[str, Any]] = []
    with sqlite3.connect(mediapark_db) as conn:
        conn.row_factory = sqlite3.Row
        for table in sample_rows:
            for column in conn.execute(f"pragma table_info({table})"):
                name = str(column["name"])
                path = f"{table}.{name}"
                desc, priority = desc_map.get(path, (f"{path} DB column.", "db_only"))
                example = sample_rows[table].get(name)
                if path in {"raw_products.in_stock", "publication_attempts.success"}:
                    data_type = "integer_boolean"
                else:
                    data_type = sqlite_type(str(column["type"]))
                rows.append(
                    make_row(
                        field_name=name,
                        field_path=path,
                        level="top_level",
                        group="db_only" if table in {"scrape_runs", "raw_products"} else "images" if table == "raw_product_images" else "specs" if table == "raw_product_specs" else "outbox",
                        data_type=data_type,
                        nullable=y(path in nullable_override or (not bool(column["notnull"]) and not bool(column["pk"]))),
                        required_for_publish="no",
                        required_for_persistence=y(bool(column["pk"]) or bool(column["notnull"])),
                        source_stage=stage_map[table],
                        source_location=f"shared/db/schema.sql::{path}",
                        store_specific_or_common="store-specific" if path in {"raw_product_specs.spec_name", "raw_product_specs.spec_value", "raw_product_specs.source_section"} else "common",
                        crm_priority=priority,
                        observed_in_sample="yes",
                        description=desc,
                        example_value=short_json(example) if isinstance(example, (dict, list)) else str(example),
                        notes=notes_map.get(path, ""),
                    )
                )

    rows.extend(
        [
            make_row(field_name="category_url_item", field_path="scrape_runs.category_urls_json[]", level="array_item", group="metadata", data_type="string", nullable="no", required_for_publish="no", required_for_persistence="no", source_stage="pipeline", source_location="shared/db/schema.sql::scrape_runs.category_urls_json", store_specific_or_common="common", crm_priority="diagnostic", observed_in_sample="yes", description="One seeded category URL inside category_urls_json.", example_value=str(run_categories[0]), notes="Useful for replay and coverage debugging."),
            make_row(field_name="stats_metric_entry", field_path="scrape_runs.stats_json.<metric_key>", level="array_item", group="metadata", data_type="dynamic key/value", nullable="yes", required_for_publish="no", required_for_persistence="no", source_stage="pipeline", source_location="shared/db/schema.sql::scrape_runs.stats_json", store_specific_or_common="common", crm_priority="diagnostic", observed_in_sample="yes", description="One run metric inside stats_json.", example_value="pages_visited_total => 28", notes="Observed stats_json includes counters like items_scraped, pages_visited_total, spec_coverage_ratio, zero_result_categories_count."),
        ]
    )
    return rows


def build_store_notes() -> list[dict[str, str]]:
    return [
        {"store_name": "mediapark", "spider_name": "mediapark", "access_mode": "HTTP-first", "listing_source_of_truth": "Anchors + escaped Next.js/inline payload product links", "pdp_source_of_truth": "JSON-LD first, then Next.js payloads, then DOM fallback", "source_id_strategy": "Trailing numeric PDP slug", "specs_strategy": "JSON/Next.js payloads + table/body fallback", "images_strategy": "JSON-LD + Next.js + og:image/img fallback", "dynamic_or_unstable_fields": "raw_specs keys vary by product/store wording", "notes": "Golden reference store. Common contract fields are stable; raw_specs labels remain dynamic."},
        {"store_name": "uzum", "spider_name": "uzum", "access_mode": "Playwright listing + HTTP-first PDP", "listing_source_of_truth": "Hydrated browser DOM snapshot injected before response serialization", "pdp_source_of_truth": "ProductGroup/Product JSON-LD first, DOM/meta fallback", "source_id_strategy": "skuId first, then numeric path suffix", "specs_strategy": "additionalProperty from ProductGroup variants, then HTML tables/itemprop", "images_strategy": "JSON-LD image arrays + meta/img fallback", "dynamic_or_unstable_fields": "raw_specs keys and some category branches drift with live taxonomy", "notes": "Main unstable area is browser-driven listing coverage, not DB contract shape."},
        {"store_name": "texnomart", "spider_name": "texnomart", "access_mode": "Browser-assisted", "listing_source_of_truth": "Anchors + serialized hrefs on /katalog/ listing pages", "pdp_source_of_truth": "JSON-LD + HTML fallbacks", "source_id_strategy": "Structured id/query/path suffix/slug fallback", "specs_strategy": "JSON-LD additionalProperty + tables/dl rows", "images_strategy": "JSON-LD + img attributes", "dynamic_or_unstable_fields": "price blocks and raw_specs labels can drift with markup changes", "notes": "Anti-bot noise exists, but stored field contract stays on the same common contour."},
        {"store_name": "alifshop", "spider_name": "alifshop", "access_mode": "HTTP-first", "listing_source_of_truth": "Moderated-offer anchors and escaped hrefs", "pdp_source_of_truth": "Meta tags + characteristics rows", "source_id_strategy": "Numeric tail in moderated-offer slug", "specs_strategy": "Characteristic rows + body fallback", "images_strategy": "og:image/twitter:image/Fortifai CDN/img attrs", "dynamic_or_unstable_fields": "external_ids key equals store name; raw_specs labels remain store-native", "notes": "Cleanest live store for CRM mapping examples, though image_urls may include noisy marketplace assets."},
    ]


def main() -> None:
    sample = load_samples()
    publish_rows = build_contract_rows(sample)
    db_rows = build_db_rows(sample)
    combined: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in publish_rows + db_rows:
        path = str(row["field_path"])
        if path in seen:
            continue
        seen.add(path)
        combined.append(row)
    store_notes = build_store_notes()

    wb = Workbook()
    ws = wb.active
    ws.title = "all_fields"
    write_sheet(ws, HEADERS, combined)
    ws = wb.create_sheet("publish_contract")
    write_sheet(ws, HEADERS, publish_rows)
    ws = wb.create_sheet("db_fields")
    write_sheet(ws, HEADERS, db_rows)
    ws = wb.create_sheet("store_notes")
    write_sheet(ws, list(store_notes[0].keys()), store_notes)
    wb.save(OUT)

    check = load_workbook(OUT, read_only=True)
    print(
        json.dumps(
            {
                "output_path": str(OUT),
                "sheets": check.sheetnames,
                "all_fields": len(combined),
                "publish_contract": len(publish_rows),
                "db_fields": len(db_rows),
                "store_notes": len(store_notes),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
