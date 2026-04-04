from __future__ import annotations

import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data" / "scraper" / "qa" / "mediapark_expanded_v1.db"
OUT_PATH = ROOT / "scraper_db_schema_for_crm.xlsx"

OVERVIEW_HEADERS = [
    "table_name",
    "purpose",
    "row_granularity",
    "created_by",
    "used_before_rabbitmq",
    "notes",
]

COLUMNS_HEADERS = [
    "table_name",
    "column_name",
    "column_order",
    "data_type",
    "nullable",
    "is_primary_key",
    "is_foreign_key",
    "default_value",
    "indexed",
    "unique",
    "references_table",
    "references_column",
    "used_for_stage",
    "description",
]

REL_HEADERS = [
    "parent_table",
    "parent_column",
    "child_table",
    "child_column",
    "relationship_type",
    "notes",
]

BOUNDARY_HEADERS = [
    "table_name",
    "boundary_role",
    "directly_used_for_publish",
    "stores_raw_scraped_data",
    "stores_metadata",
    "stores_publication_state",
    "notes",
]

CRM_HEADERS = ["topic", "value", "notes"]


def y(value: bool) -> str:
    return "yes" if value else "no"


def sqlite_type(raw_type: str) -> str:
    normalized = (raw_type or "").strip().upper()
    if normalized in {"TEXT", "INTEGER", "REAL", "BLOB", "NUMERIC"}:
        return normalized.lower()
    if not normalized:
        return "text"
    return normalized.lower()


TABLE_META: dict[str, dict[str, str]] = {
    "schema_migrations": {
        "purpose": "Tracks applied scraper DB schema versions.",
        "row_granularity": "one row per applied migration version",
        "created_by": "SQLiteScraperStore.ensure_schema",
        "used_before_rabbitmq": "yes",
        "notes": "Infrastructure/support table; not product data.",
        "boundary_role": "schema_bootstrap",
        "directly_used_for_publish": "no",
        "stores_raw_scraped_data": "no",
        "stores_metadata": "yes",
        "stores_publication_state": "no",
    },
    "scrape_runs": {
        "purpose": "Tracks one scraper run per spider/store with counters, seeded categories, and run stats.",
        "row_granularity": "one row per scraper run",
        "created_by": "ScraperStoragePipeline.open_spider -> ScraperPersistenceService.start_run",
        "used_before_rabbitmq": "yes",
        "notes": "Main run/audit table. Useful for QA, coverage, replay, and XLSX exports.",
        "boundary_role": "run_tracking",
        "directly_used_for_publish": "indirectly",
        "stores_raw_scraped_data": "no",
        "stores_metadata": "yes",
        "stores_publication_state": "no",
    },
    "raw_products": {
        "purpose": "Stores minimally structured raw product snapshots before RabbitMQ publication.",
        "row_granularity": "one row per scrape_run_id + identity_key",
        "created_by": "ScraperStoragePipeline.process_item -> SQLiteScraperStore.persist_snapshot",
        "used_before_rabbitmq": "yes",
        "notes": "Primary product table for pre-RabbitMQ storage, replay, audit, and XLSX export.",
        "boundary_role": "raw_product_storage",
        "directly_used_for_publish": "yes",
        "stores_raw_scraped_data": "yes",
        "stores_metadata": "yes",
        "stores_publication_state": "yes",
    },
    "raw_product_images": {
        "purpose": "Stores product images as first-class rows linked to raw_products.",
        "row_granularity": "one row per raw product image",
        "created_by": "SQLiteScraperStore._replace_child_images",
        "used_before_rabbitmq": "yes",
        "notes": "Useful for quality checks and SQL/XLSX exports without parsing JSON arrays.",
        "boundary_role": "raw_product_assets",
        "directly_used_for_publish": "indirectly",
        "stores_raw_scraped_data": "yes",
        "stores_metadata": "no",
        "stores_publication_state": "no",
    },
    "raw_product_specs": {
        "purpose": "Stores flattened raw specs linked to raw_products.",
        "row_granularity": "one row per flattened spec entry",
        "created_by": "SQLiteScraperStore._replace_child_specs",
        "used_before_rabbitmq": "yes",
        "notes": "Useful for quality checks and SQL/XLSX exports without parsing JSON maps.",
        "boundary_role": "raw_product_specs",
        "directly_used_for_publish": "indirectly",
        "stores_raw_scraped_data": "yes",
        "stores_metadata": "no",
        "stores_publication_state": "no",
    },
    "publication_outbox": {
        "purpose": "Persistent outbox queue that holds one publishable event per raw product snapshot.",
        "row_granularity": "one row per raw product snapshot / event",
        "created_by": "SQLiteScraperStore._upsert_outbox_row",
        "used_before_rabbitmq": "yes",
        "notes": "Core boundary table between scraper persistence and publisher service.",
        "boundary_role": "rabbitmq_publication_boundary",
        "directly_used_for_publish": "yes",
        "stores_raw_scraped_data": "no",
        "stores_metadata": "yes",
        "stores_publication_state": "yes",
    },
    "publication_attempts": {
        "purpose": "Audit log of publisher attempts for each outbox row.",
        "row_granularity": "one row per publish attempt",
        "created_by": "Publisher service -> mark_outbox_published / mark_outbox_failed",
        "used_before_rabbitmq": "yes",
        "notes": "Publisher-side audit table. Important for retry/debug, not for product payload itself.",
        "boundary_role": "publication_audit",
        "directly_used_for_publish": "yes",
        "stores_raw_scraped_data": "no",
        "stores_metadata": "yes",
        "stores_publication_state": "yes",
    },
}

STAGE_BY_TABLE = {
    "schema_migrations": "schema bootstrap",
    "scrape_runs": "scrape run tracking",
    "raw_products": "raw product storage",
    "raw_product_images": "images storage",
    "raw_product_specs": "specs storage",
    "publication_outbox": "outbox",
    "publication_attempts": "publication attempts",
}

COLUMN_DESCRIPTIONS: dict[tuple[str, str], str] = {
    ("schema_migrations", "version"): "Migration version identifier.",
    ("schema_migrations", "applied_at"): "When the migration version was applied.",
    ("scrape_runs", "id"): "Surrogate primary key for scrape run row.",
    ("scrape_runs", "run_id"): "Stable run identifier shared by all items in one crawl.",
    ("scrape_runs", "store_name"): "Store name for the run.",
    ("scrape_runs", "spider_name"): "Scrapy spider name that created the run.",
    ("scrape_runs", "started_at"): "Run start timestamp.",
    ("scrape_runs", "finished_at"): "Run finish timestamp.",
    ("scrape_runs", "status"): "Run status such as completed, partial_failure, or failed.",
    ("scrape_runs", "items_scraped"): "Number of items yielded by spider in this run.",
    ("scrape_runs", "items_persisted"): "Number of items persisted into raw_products in this run.",
    ("scrape_runs", "items_failed"): "Number of failed or dropped items in this run.",
    ("scrape_runs", "category_urls_json"): "JSON array of seeded category URLs used to start the run.",
    ("scrape_runs", "stats_json"): "JSON map of crawl/run metrics, coverage counters, and QA stats.",
    ("scrape_runs", "created_at"): "Row creation timestamp.",
    ("scrape_runs", "updated_at"): "Row update timestamp.",
    ("raw_products", "id"): "Surrogate primary key for raw product row.",
    ("raw_products", "scrape_run_id"): "Run id linking product snapshot back to scrape_runs.run_id.",
    ("raw_products", "store_name"): "Store name for the product snapshot.",
    ("raw_products", "source_id"): "Store-native product id if extracted.",
    ("raw_products", "source_url"): "Canonical source PDP URL.",
    ("raw_products", "identity_key"): "Stable scraper identity key used for dedupe and CRM upsert identity.",
    ("raw_products", "title"): "Persisted minimally structured title.",
    ("raw_products", "brand"): "Persisted brand if available.",
    ("raw_products", "price_raw"): "Persisted raw price string as scraped from the store.",
    ("raw_products", "in_stock"): "Persisted nullable stock flag stored as integer 0/1 in SQLite.",
    ("raw_products", "description"): "Persisted store-native description if available.",
    ("raw_products", "category_hint"): "Persisted lightweight category hint.",
    ("raw_products", "external_ids_json"): "JSON object with external ids keyed by source/store.",
    ("raw_products", "payload_hash"): "Business fingerprint of the minimally structured payload.",
    ("raw_products", "raw_payload_json"): "JSON snapshot of the validated raw spider item.",
    ("raw_products", "structured_payload_json"): "JSON snapshot of the shaped payload that also feeds the outbox event.",
    ("raw_products", "scraped_at"): "UTC timestamp when the product snapshot was created.",
    ("raw_products", "publication_state"): "Current publication lifecycle state mirrored from outbox transitions.",
    ("raw_products", "created_at"): "Row creation timestamp.",
    ("raw_products", "updated_at"): "Row update timestamp.",
    ("raw_product_images", "id"): "Surrogate primary key for image row.",
    ("raw_product_images", "raw_product_id"): "Foreign key to raw_products.id.",
    ("raw_product_images", "image_url"): "One image URL for the product.",
    ("raw_product_images", "position"): "Ordinal position of image inside the product image list.",
    ("raw_product_images", "created_at"): "Row creation timestamp.",
    ("raw_product_images", "updated_at"): "Row update timestamp.",
    ("raw_product_specs", "id"): "Surrogate primary key for spec row.",
    ("raw_product_specs", "raw_product_id"): "Foreign key to raw_products.id.",
    ("raw_product_specs", "spec_name"): "Flattened raw spec label.",
    ("raw_product_specs", "spec_value"): "Flattened raw spec value.",
    ("raw_product_specs", "source_section"): "Optional parent section when source raw_specs contained nested groups.",
    ("raw_product_specs", "position"): "Ordinal position of flattened spec row.",
    ("raw_product_specs", "created_at"): "Row creation timestamp.",
    ("raw_product_specs", "updated_at"): "Row update timestamp.",
    ("publication_outbox", "id"): "Surrogate primary key for outbox row.",
    ("publication_outbox", "raw_product_id"): "Unique link to raw_products.id for the source snapshot.",
    ("publication_outbox", "event_id"): "Stable event id for the payload published to RabbitMQ.",
    ("publication_outbox", "event_type"): "Logical event type stored with the outbox row.",
    ("publication_outbox", "schema_version"): "Schema version of the payload contract.",
    ("publication_outbox", "scrape_run_id"): "Run id linking the outbox row back to scrape_runs.run_id.",
    ("publication_outbox", "store_name"): "Store name copied into the outbox row.",
    ("publication_outbox", "source_id"): "Optional store-native product id copied into the outbox row.",
    ("publication_outbox", "source_url"): "Source PDP URL copied into the outbox row.",
    ("publication_outbox", "payload_hash"): "Business fingerprint copied into the outbox row.",
    ("publication_outbox", "exchange_name"): "RabbitMQ exchange configured for publication.",
    ("publication_outbox", "routing_key"): "RabbitMQ routing key configured for publication.",
    ("publication_outbox", "payload_json"): "Serialized JSON event payload to be published by the publisher service.",
    ("publication_outbox", "status"): "Outbox status such as pending, publishing, published, retryable, or failed.",
    ("publication_outbox", "available_at"): "When the outbox row becomes eligible for claim/retry.",
    ("publication_outbox", "published_at"): "When the publisher service marked the outbox row as published.",
    ("publication_outbox", "retry_count"): "Number of publication attempts already consumed by this row.",
    ("publication_outbox", "last_error"): "Last publication error message if the row failed or is retryable.",
    ("publication_outbox", "lease_owner"): "Publisher service lease owner while the row is being published.",
    ("publication_outbox", "lease_expires_at"): "Lease expiry timestamp for in-progress outbox rows.",
    ("publication_outbox", "created_at"): "Row creation timestamp.",
    ("publication_outbox", "updated_at"): "Row update timestamp.",
    ("publication_attempts", "id"): "Surrogate primary key for publication attempt row.",
    ("publication_attempts", "outbox_id"): "Foreign key to publication_outbox.id.",
    ("publication_attempts", "attempt_number"): "1-based publication attempt number for the outbox row.",
    ("publication_attempts", "attempted_at"): "When the publisher attempted delivery.",
    ("publication_attempts", "success"): "Whether the attempt succeeded (1) or failed (0).",
    ("publication_attempts", "error_message"): "Error text for failed publication attempts.",
    ("publication_attempts", "publisher_name"): "Publisher service name that wrote the attempt.",
    ("publication_attempts", "created_at"): "Row creation timestamp.",
}


def load_schema(conn: sqlite3.Connection) -> list[str]:
    rows = conn.execute(
        "select name from sqlite_master where type = 'table' and name not like 'sqlite_%' order by name"
    ).fetchall()
    return [str(row[0]) for row in rows]


def load_index_metadata(conn: sqlite3.Connection, table: str) -> tuple[set[str], set[str], dict[str, str]]:
    indexed_cols: set[str] = set()
    unique_cols: set[str] = set()
    index_notes: dict[str, str] = {}
    index_list = conn.execute(f"pragma index_list({table})").fetchall()
    for idx in index_list:
        index_name = str(idx[1])
        is_unique = bool(idx[2])
        cols = [str(row[2]) for row in conn.execute(f"pragma index_info({index_name})").fetchall()]
        for col in cols:
            indexed_cols.add(col)
            if is_unique:
                unique_cols.add(col)
        if cols:
            joined = ", ".join(cols)
            index_notes[index_name] = f"cols=({joined}), unique={y(is_unique)}"
    return indexed_cols, unique_cols, index_notes


def build_overview_rows(tables: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for table in tables:
        meta = TABLE_META.get(table, {})
        rows.append(
            {
                "table_name": table,
                "purpose": meta.get("purpose", f"{table} scraper-side table."),
                "row_granularity": meta.get("row_granularity", ""),
                "created_by": meta.get("created_by", ""),
                "used_before_rabbitmq": meta.get("used_before_rabbitmq", "yes"),
                "notes": meta.get("notes", ""),
            }
        )
    return rows


def build_columns_rows(conn: sqlite3.Connection, tables: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for table in tables:
        indexed_cols, unique_cols, _ = load_index_metadata(conn, table)
        fks = {
            str(fk["from"]): (str(fk["table"]), str(fk["to"]))
            for fk in conn.execute(f"pragma foreign_key_list({table})").fetchall()
        }
        for column in conn.execute(f"pragma table_info({table})").fetchall():
            column_name = str(column["name"])
            ref_table, ref_col = fks.get(column_name, ("", ""))
            rows.append(
                {
                    "table_name": table,
                    "column_name": column_name,
                    "column_order": str(int(column["cid"]) + 1),
                    "data_type": sqlite_type(str(column["type"])),
                    "nullable": y(not bool(column["notnull"]) and not bool(column["pk"])),
                    "is_primary_key": y(bool(column["pk"])),
                    "is_foreign_key": y(column_name in fks),
                    "default_value": "" if column["dflt_value"] is None else str(column["dflt_value"]),
                    "indexed": y(column_name in indexed_cols),
                    "unique": y(column_name in unique_cols or bool(column["pk"])),
                    "references_table": ref_table,
                    "references_column": ref_col,
                    "used_for_stage": STAGE_BY_TABLE.get(table, ""),
                    "description": COLUMN_DESCRIPTIONS.get((table, column_name), f"{table}.{column_name} column."),
                }
            )
    return rows


def build_relationship_rows(conn: sqlite3.Connection, tables: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for table in tables:
        for fk in conn.execute(f"pragma foreign_key_list({table})").fetchall():
            parent_table = str(fk["table"])
            parent_column = str(fk["to"])
            child_table = table
            child_column = str(fk["from"])
            rows.append(
                {
                    "parent_table": parent_table,
                    "parent_column": parent_column,
                    "child_table": child_table,
                    "child_column": child_column,
                    "relationship_type": "one-to-many",
                    "notes": "ON DELETE CASCADE" if str(fk["on_delete"]).upper() == "CASCADE" else "",
                }
            )
    return rows


def build_boundary_rows(tables: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for table in tables:
        meta = TABLE_META.get(table, {})
        rows.append(
            {
                "table_name": table,
                "boundary_role": meta.get("boundary_role", ""),
                "directly_used_for_publish": meta.get("directly_used_for_publish", "no"),
                "stores_raw_scraped_data": meta.get("stores_raw_scraped_data", "no"),
                "stores_metadata": meta.get("stores_metadata", "no"),
                "stores_publication_state": meta.get("stores_publication_state", "no"),
                "notes": meta.get("notes", ""),
            }
        )
    return rows


def build_crm_notes() -> list[dict[str, str]]:
    return [
        {
            "topic": "RabbitMQ producer source",
            "value": "publication_outbox.payload_json",
            "notes": "Publisher service claims publication_outbox rows and publishes payload_json to RabbitMQ.",
        },
        {
            "topic": "Primary raw product table",
            "value": "raw_products",
            "notes": "Main pre-RabbitMQ product snapshot table. Contains identity_key, payload_hash, structured/raw payload JSON, and publication_state.",
        },
        {
            "topic": "Specs storage",
            "value": "raw_product_specs",
            "notes": "Flattened specs table for SQL/XLSX export. Full raw spec blob also stays in raw_products.structured_payload_json and raw_payload_json.",
        },
        {
            "topic": "Images storage",
            "value": "raw_product_images",
            "notes": "One row per image URL with position. Easier for analytics and export than parsing JSON arrays.",
        },
        {
            "topic": "Run tracking",
            "value": "scrape_runs",
            "notes": "Run-level counters, seeded categories, and stats_json. Useful for QA, audit, and export slicing.",
        },
        {
            "topic": "Publication state",
            "value": "publication_outbox + raw_products.publication_state + publication_attempts",
            "notes": "publication_outbox stores the publish queue state; raw_products mirrors publication_state; publication_attempts stores retry/audit history.",
        },
        {
            "topic": "Boundary before RabbitMQ",
            "value": "Everything in scraper DB is pre-RabbitMQ state",
            "notes": "scraper writes rows into raw_products/images/specs and publication_outbox first; publisher service reads the outbox afterward.",
        },
        {
            "topic": "Best tables for future XLSX export",
            "value": "scrape_runs, raw_products, raw_product_images, raw_product_specs",
            "notes": "These four tables provide run context plus tabular product/spec/image data without parsing publisher/audit tables.",
        },
        {
            "topic": "Service/bootstrap table",
            "value": "schema_migrations",
            "notes": "Not product data, but included because it is a real scraper DB table discovered by introspection.",
        },
    ]


def format_sheet(ws, headers: list[str], rows: list[dict[str, str]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    wrap_cols = {
        "purpose",
        "notes",
        "description",
        "created_by",
    }
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        width = max(len(header), *(len(str(row.get(header, ""))) for row in rows)) if rows else len(header)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 12), 60)
        for cell in ws[get_column_letter(idx)]:
            cell.alignment = Alignment(vertical="top", wrap_text=header in wrap_cols)


def main() -> None:
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        tables = load_schema(conn)
        overview = build_overview_rows(tables)
        columns = build_columns_rows(conn, tables)
        relationships = build_relationship_rows(conn, tables)
        boundary = build_boundary_rows(tables)
        crm_notes = build_crm_notes()

    wb = Workbook()
    ws = wb.active
    ws.title = "tables_overview"
    format_sheet(ws, OVERVIEW_HEADERS, overview)

    ws = wb.create_sheet("columns_all")
    format_sheet(ws, COLUMNS_HEADERS, columns)

    ws = wb.create_sheet("relationships")
    format_sheet(ws, REL_HEADERS, relationships)

    ws = wb.create_sheet("rabbitmq_boundary")
    format_sheet(ws, BOUNDARY_HEADERS, boundary)

    ws = wb.create_sheet("crm_notes")
    format_sheet(ws, CRM_HEADERS, crm_notes)

    wb.save(OUT_PATH)

    check = load_workbook(OUT_PATH, read_only=True)
    print(
        {
            "output_path": str(OUT_PATH),
            "sheets": check.sheetnames,
            "tables": len(tables),
            "columns": len(columns),
            "relationships": len(relationships),
        }
    )


if __name__ == "__main__":
    main()
