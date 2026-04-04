from __future__ import annotations

import sqlite3
from datetime import datetime, UTC
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DB_DIR = ROOT / "data" / "scraper" / "qa"
OUT_PATH = ROOT / "scraper_qa_databases_export.xlsx"

TABLES = [
    "schema_migrations",
    "scrape_runs",
    "raw_products",
    "raw_product_images",
    "raw_product_specs",
    "publication_outbox",
    "publication_attempts",
]


def format_sheet(ws) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, column in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 60)
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def list_tables(conn: sqlite3.Connection) -> set[str]:
    rows = conn.execute(
        "select name from sqlite_master where type = 'table' and name not like 'sqlite_%'"
    ).fetchall()
    return {str(row[0]) for row in rows}


def table_columns(conn: sqlite3.Connection, table: str) -> list[str]:
    rows = conn.execute(f"pragma table_info({table})").fetchall()
    return [str(row[1]) for row in rows]


def main() -> None:
    db_paths = sorted(DB_DIR.glob("*.db"))
    if not db_paths:
        raise SystemExit(f"No SQLite DB files found in {DB_DIR}")

    summary_headers = [
        "source_db",
        "table_name",
        "row_count",
        "db_size_bytes",
        "last_modified",
    ]
    summary_rows: list[list[object]] = []
    table_rows: dict[str, list[list[object]]] = {table: [] for table in TABLES}
    table_headers: dict[str, list[str]] = {}

    for db_path in db_paths:
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            existing_tables = list_tables(conn)
            for table in TABLES:
                if table not in existing_tables:
                    continue
                columns = table_columns(conn, table)
                headers = ["source_db"] + columns
                table_headers.setdefault(table, headers)
                count = conn.execute(f"select count(*) from {table}").fetchone()[0]
                summary_rows.append(
                    [
                        db_path.name,
                        table,
                        count,
                        db_path.stat().st_size,
                        datetime.fromtimestamp(db_path.stat().st_mtime, UTC).isoformat(),
                    ]
                )
                rows = conn.execute(f"select * from {table}").fetchall()
                for row in rows:
                    table_rows[table].append([db_path.name] + [row[column] for column in columns])

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(summary_headers)
    for row in summary_rows:
        ws.append(row)
    format_sheet(ws)

    for table in TABLES:
        headers = table_headers.get(table)
        if not headers:
            continue
        ws = wb.create_sheet(table[:31])
        ws.append(headers)
        for row in table_rows[table]:
            ws.append(row)
        format_sheet(ws)

    wb.save(OUT_PATH)
    check = load_workbook(OUT_PATH, read_only=True)
    print(
        {
            "output_path": str(OUT_PATH),
            "sheets": check.sheetnames,
            "db_files": len(db_paths),
            "summary_rows": len(summary_rows),
            "exported_tables": len(check.sheetnames) - 1,
        }
    )


if __name__ == "__main__":
    main()
